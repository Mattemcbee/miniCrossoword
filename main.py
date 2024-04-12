import re
from openpyxl import Workbook

def extractMessage(py_file_path):
    try:
        with open(py_file_path, 'r', encoding='utf-8') as file:
            text = file.read()
            return text
    except Exception as e:
        return f"Error: {e}"

python_file_name = 'data.py'
result = extractMessage(python_file_name)

crossRegex = re.compile(r'\d:\d{2}')  
dateRegex = re.compile(r'\d*/\d*/\d{4}')  
nameRegex = re.compile(r'([a-zA-Z]+):') 

# Find all names, dates, and times
extractedNames = nameRegex.findall(result)
extractedTimes = crossRegex.findall(result)
extractedDates = dateRegex.findall(result)

# Combine names, dates, and times into tuples
data = list(zip(extractedNames, extractedDates, extractedTimes))

# Organize data by date and then by name
scores_dict = {}
for name, date, time in data:
    if date not in scores_dict:
        scores_dict[date] = {}
    if name not in scores_dict[date]:
        scores_dict[date][name] = []
    scores_dict[date][name].append(time)

# Sort scores_dict by date
scores_dict_sorted = dict(sorted(scores_dict.items()))

# Create a new workbook
wb = Workbook()
ws = wb.active

# Write headers
ws['A1'] = 'Date'
unique_names = sorted(set(extractedNames))  # Get sorted unique names
for col, name in enumerate(unique_names, start=2):
    ws.cell(row=1, column=col, value=name)

# Write data to the workbook
for row, (date, scores_by_name) in enumerate(scores_dict_sorted.items(), start=2):
    ws.cell(row=row, column=1, value=date)
    for col, name in enumerate(unique_names, start=2):
        scores = scores_by_name.get(name, [])
        ws.cell(row=row, column=col, value=', '.join(scores))

# Save the workbook
excel_file_name = 'scores.xlsx'
try:
    wb.save(excel_file_name)
    print(f"Excel file '{excel_file_name}' has been created.")
except Exception as e:
    print(f"Error saving Excel file: {e}")
finally:
    # Close the workbook
    wb.close()
